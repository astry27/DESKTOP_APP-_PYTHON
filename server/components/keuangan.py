# Path: server/components/keuangan.py

from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
                             QTableWidgetItem, QLabel, QPushButton, QHeaderView,
                             QMessageBox, QLineEdit, QComboBox, QAbstractItemView,
                             QTabWidget, QDialog, QFormLayout, QDateEdit, QSpinBox,
                             QTextEdit, QFileDialog, QFrame)
from PyQt5.QtCore import Qt, pyqtSignal, QSize, QRect, QTimer, QDate
from PyQt5.QtGui import QFont, QColor, QPainter, QBrush
from database import DatabaseManager
import datetime

class WordWrapHeaderView(QHeaderView):
    """Custom header view with word wrap and center alignment support"""
    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
        self.setDefaultAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        self.setSectionsClickable(True)
        self.setHighlightSections(True)

    def sectionSizeFromContents(self, logicalIndex):
        """Calculate section size based on wrapped text"""
        if self.model():
            text = self.model().headerData(logicalIndex, self.orientation(), Qt.DisplayRole)
            if text:
                width = self.sectionSize(logicalIndex)
                if width <= 0:
                    width = self.defaultSectionSize()
                font = self.font()
                font.setBold(True)
                fm = self.fontMetrics()
                rect = fm.boundingRect(0, 0, width - 8, 1000,
                                      Qt.AlignCenter | Qt.TextWordWrap, str(text))
                return QSize(width, max(rect.height() + 12, 25))
        return super().sectionSizeFromContents(logicalIndex)

    def paintSection(self, painter, rect, logicalIndex):
        """Paint section with word wrap and center alignment"""
        painter.save()
        bg_color = QColor(242, 242, 242)
        painter.fillRect(rect, bg_color)
        border_color = QColor(212, 212, 212)
        painter.setPen(border_color)
        painter.drawLine(rect.topRight(), rect.bottomRight())
        painter.drawLine(rect.bottomLeft(), rect.bottomRight())

        if self.model():
            text = self.model().headerData(logicalIndex, self.orientation(), Qt.DisplayRole)
            if text:
                font = self.font()
                font.setBold(True)
                painter.setFont(font)
                text_color = QColor(51, 51, 51)
                painter.setPen(text_color)
                text_rect = rect.adjusted(4, 4, -4, -4)
                painter.drawText(text_rect, Qt.AlignCenter | Qt.TextWordWrap, str(text))
        painter.restore()

class KeuanganDialog(QDialog):
    """Dialog untuk tambah/edit transaksi keuangan kategorial"""
    def __init__(self, parent=None, keuangan_data=None):
        super().__init__(parent)
        self.keuangan_data = keuangan_data if keuangan_data else {}
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("Tambah/Edit Transaksi Keuangan")
        self.setModal(True)
        self.resize(500, 400)

        layout = QVBoxLayout(self)
        form_layout = QFormLayout()
        form_layout.setVerticalSpacing(8)

        # Tanggal
        self.tanggal = QDateEdit()
        self.tanggal.setCalendarPopup(True)
        if self.keuangan_data.get('tanggal'):
            try:
                date_obj = datetime.datetime.strptime(self.keuangan_data['tanggal'], '%Y-%m-%d').date()
                self.tanggal.setDate(QDate.fromString(date_obj.isoformat(), Qt.ISODate))
            except:
                self.tanggal.setDate(QDate.currentDate())
        else:
            self.tanggal.setDate(QDate.currentDate())
        form_layout.addRow("Tanggal:", self.tanggal)

        # Jenis (Pemasukan/Pengeluaran) - mapping ke field 'jenis' di database
        self.jenis = QComboBox()
        self.jenis.addItems(['Pemasukan', 'Pengeluaran'])
        if self.keuangan_data.get('jenis'):
            index = self.jenis.findText(self.keuangan_data['jenis'])
            if index >= 0:
                self.jenis.setCurrentIndex(index)
        form_layout.addRow("Jenis Transaksi:", self.jenis)

        # Kategori - mapping ke field 'kategori' di database
        self.kategori = QComboBox()
        self.kategori.setEditable(True)
        self.kategori.addItems([
            "Kolekte Mingguan", "Persembahan", "Donasi", "Operasional",
            "Pemeliharaan Gedung", "Program Gereja", "Lainnya"
        ])
        if self.keuangan_data.get('kategori'):
            index = self.kategori.findText(self.keuangan_data['kategori'])
            if index >= 0:
                self.kategori.setCurrentIndex(index)
            else:
                self.kategori.setCurrentText(self.keuangan_data['kategori'])
        form_layout.addRow("Kategori:", self.kategori)

        # Keterangan
        self.keterangan = QTextEdit()
        self.keterangan.setMaximumHeight(80)
        self.keterangan.setText(self.keuangan_data.get('keterangan', ''))
        form_layout.addRow("Keterangan:", self.keterangan)

        # Jumlah
        self.jumlah = QSpinBox()
        self.jumlah.setRange(0, 999999999)
        self.jumlah.setValue(int(self.keuangan_data.get('jumlah', 0)))
        self.jumlah.setSuffix(" Rupiah")
        form_layout.addRow("Jumlah:", self.jumlah)

        layout.addLayout(form_layout)

        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        save_button = QPushButton("Simpan")
        save_button.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                padding: 8px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2ecc71;
            }
        """)
        save_button.clicked.connect(self.accept)

        cancel_button = QPushButton("Batal")
        cancel_button.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                padding: 8px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #7f8c8d;
            }
        """)
        cancel_button.clicked.connect(self.reject)

        button_layout.addWidget(save_button)
        button_layout.addWidget(cancel_button)
        layout.addLayout(button_layout)

    def get_data(self):
        return {
            'tanggal': self.tanggal.date().toString(Qt.ISODate),
            'jenis': self.jenis.currentText(),  # Pemasukan/Pengeluaran
            'kategori': self.kategori.currentText(),  # Kolekte, Donasi, dll
            'keterangan': self.keterangan.toPlainText(),
            'jumlah': self.jumlah.value(),
        }

class KeuanganWRWidget(QWidget):
    """Tab WR - Keuangan dari input WR/client"""
    log_message = pyqtSignal(str)

    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.keuangan_data = []
        self.filtered_data = []
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Header
        header_layout = QHBoxLayout()
        title = QLabel("Data Keuangan WR")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        header_layout.addWidget(title)
        header_layout.addStretch()

        refresh_btn = QPushButton("Refresh")
        refresh_btn.clicked.connect(self.load_data)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                padding: 6px 12px;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        header_layout.addWidget(refresh_btn)
        layout.addLayout(header_layout)

        # Filter
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Cari:"))
        self.search_input = QLineEdit(self)
        self.search_input.setPlaceholderText("Cari keterangan...")
        self.search_input.textChanged.connect(self.filter_data)
        filter_layout.addWidget(self.search_input)

        filter_layout.addWidget(QLabel("Jenis:"))
        self.jenis_filter = QComboBox(self)
        self.jenis_filter.addItems(["Semua", "Pemasukan", "Pengeluaran"])
        self.jenis_filter.currentTextChanged.connect(self.filter_data)
        filter_layout.addWidget(self.jenis_filter)

        filter_layout.addWidget(QLabel("User:"))
        self.user_filter = QComboBox(self)
        self.user_filter.addItem("Semua User")
        self.user_filter.currentTextChanged.connect(self.filter_data)
        filter_layout.addWidget(self.user_filter)

        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(8)

        custom_header = WordWrapHeaderView(Qt.Horizontal, self.table)
        self.table.setHorizontalHeader(custom_header)

        self.table.setHorizontalHeaderLabels([
            "Tanggal", "Jenis", "Kategori", "Keterangan",
            "Jumlah", "Aksi"
        ])

        self.apply_professional_table_style(self.table)

        self.table.setColumnWidth(0, 110)
        self.table.setColumnWidth(1, 100)
        self.table.setColumnWidth(2, 140)
        self.table.setColumnWidth(3, 180)
        self.table.setColumnWidth(4, 110)
        self.table.setColumnWidth(5, 110)
        self.table.setColumnWidth(6, 120)
        self.table.setColumnWidth(7, 130)
        self.table.setColumnWidth(8, 100)

        header = self.table.horizontalHeader()
        header.sectionResized.connect(self.update_header_height)

        QTimer.singleShot(100, lambda: self.update_header_height(0, 0, 0))

        layout.addWidget(self.table)

        # Ringkasan Keuangan
        stats_frame = QLabel("Ringkasan Keuangan:")
        stats_frame.setFont(QFont("Arial", 10, QFont.Bold))
        layout.addWidget(stats_frame)

        self.stats_layout = QHBoxLayout()

        self.total_label = QLabel("Total Transaksi: 0")
        self.total_label.setFont(QFont("Arial", 9))
        self.stats_layout.addWidget(self.total_label)

        self.stats_layout.addWidget(QLabel("|"))

        self.pemasukan_label = QLabel("Total Pemasukan: Rp 0")
        self.pemasukan_label.setFont(QFont("Arial", 9, QFont.Bold))
        self.pemasukan_label.setStyleSheet("color: #27ae60;")
        self.stats_layout.addWidget(self.pemasukan_label)

        self.stats_layout.addWidget(QLabel("|"))

        self.pengeluaran_label = QLabel("Total Pengeluaran: Rp 0")
        self.pengeluaran_label.setFont(QFont("Arial", 9, QFont.Bold))
        self.pengeluaran_label.setStyleSheet("color: #e74c3c;")
        self.stats_layout.addWidget(self.pengeluaran_label)

        self.stats_layout.addWidget(QLabel("|"))

        self.saldo_label = QLabel("Saldo: Rp 0")
        self.saldo_label.setFont(QFont("Arial", 10, QFont.Bold))
        self.stats_layout.addWidget(self.saldo_label)

        self.stats_layout.addStretch()
        layout.addLayout(self.stats_layout)

    def update_header_height(self, logicalIndex, oldSize, newSize):
        if hasattr(self, 'table'):
            header = self.table.horizontalHeader()
            header.setMinimumHeight(25)
            max_height = 25
            for i in range(header.count()):
                size = header.sectionSizeFromContents(i)
                max_height = max(max_height, size.height())
            header.setFixedHeight(max_height)

    def apply_professional_table_style(self, table):
        header_font = QFont()
        header_font.setBold(True)
        header_font.setPointSize(9)
        table.horizontalHeader().setFont(header_font)

        table.horizontalHeader().setStyleSheet("""
            QHeaderView::section {
                background-color: #f2f2f2;
                border: none;
                border-bottom: 1px solid #d4d4d4;
                border-right: 1px solid #d4d4d4;
                padding: 6px 4px;
                font-weight: bold;
                color: #333333;
            }
        """)

        header = table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        header.setSectionsClickable(True)
        header.setMinimumHeight(25)
        header.setMinimumSectionSize(50)
        header.setMaximumSectionSize(500)
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setStretchLastSection(True)

        table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d4d4d4;
                background-color: white;
                border: 1px solid #d4d4d4;
                selection-background-color: #cce7ff;
                font-family: 'Calibri', 'Segoe UI', Arial, sans-serif;
                font-size: 9pt;
                outline: none;
            }
            QTableWidget::item {
                border: none;
                padding: 4px 6px;
                min-height: 18px;
            }
            QTableWidget::item:selected {
                background-color: #cce7ff;
                color: black;
            }
            QTableWidget::item:focus {
                border: 2px solid #0078d4;
            }
        """)

        table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)

        table.verticalHeader().setDefaultSectionSize(20)
        table.setSelectionBehavior(QAbstractItemView.SelectItems)
        table.setAlternatingRowColors(False)
        table.verticalHeader().setVisible(True)
        table.verticalHeader().setStyleSheet("""
            QHeaderView::section {
                background-color: #f2f2f2;
                border: none;
                border-bottom: 1px solid #d4d4d4;
                border-right: 1px solid #d4d4d4;
                padding: 2px;
                font-weight: normal;
                color: #333333;
                text-align: center;
                width: 30px;
            }
        """)

        table.setShowGrid(True)
        table.setGridStyle(Qt.SolidLine)
        table.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.EditKeyPressed)
        table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        table.setMinimumHeight(150)
        table.setSizeAdjustPolicy(QAbstractItemView.AdjustToContents)

    def load_data(self):
        try:
            success, data = self.db_manager.get_keuangan_list()
            if success:
                self.keuangan_data = data if isinstance(data, list) else []
                # Sort data by ID ascending (terlama ke terbaru)
                self.keuangan_data = sorted(self.keuangan_data, key=lambda x: x.get('id', 0), reverse=False)
                self.filtered_data = self.keuangan_data.copy()
                self.update_user_filter()
                self.update_table()
                self.update_stats()
                self.log_message.emit(f"Data keuangan WR dimuat: {len(self.keuangan_data)} record (diurutkan dari terlama)")
            else:
                QMessageBox.critical(self, "Error", f"Gagal memuat data: {data}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat data: {str(e)}")

    def update_user_filter(self):
        current_text = self.user_filter.currentText()
        self.user_filter.clear()
        self.user_filter.addItem("Semua User")

        users = set()
        for item in self.keuangan_data:
            user_name = item.get('nama_user') or item.get('dibuat_oleh') or item.get('username') or ''
            if user_name and user_name != 'System':
                users.add(user_name)

        for user in sorted(users):
            self.user_filter.addItem(user)

        index = self.user_filter.findText(current_text)
        if index >= 0:
            self.user_filter.setCurrentIndex(index)

    def filter_data(self):
        search = self.search_input.text().lower()
        jenis = self.jenis_filter.currentText()
        user_filter = self.user_filter.currentText()

        self.filtered_data = []
        for item in self.keuangan_data:
            if search and search not in str(item.get('keterangan', '')).lower():
                continue

            if jenis != "Semua" and item.get('kategori') != jenis:
                continue

            if user_filter != "Semua User":
                user_name = item.get('nama_user') or item.get('dibuat_oleh') or item.get('username') or ''
                if user_name != user_filter:
                    continue

            self.filtered_data.append(item)

        # Sort filtered data by ID ascending (terlama ke terbaru)
        self.filtered_data = sorted(self.filtered_data, key=lambda x: x.get('id', 0), reverse=False)

        self.update_table()
        self.update_stats()

    def update_table(self):
        self.table.setRowCount(len(self.filtered_data))

        # Calculate running balance (saldo)
        running_balance = 0.0

        for row, item in enumerate(self.filtered_data):
            # Format tanggal transaksi - dd/mm/yyyy saja tanpa hari
            tanggal = item.get('tanggal', '')
            if tanggal:
                try:
                    if isinstance(tanggal, str):
                        tanggal_str = tanggal.strip()

                        # Handle RFC 1123 / GMT format (Mon, 17 Nov 2025 00:00:00 GMT)
                        if 'GMT' in tanggal_str or tanggal_str.count(',') == 1:
                            try:
                                # Parse RFC 1123 format
                                from email.utils import parsedate_to_datetime
                                date_obj = parsedate_to_datetime(tanggal_str)
                            except:
                                # Fallback: extract date parts manually
                                # Format: "Mon, 17 Nov 2025 00:00:00 GMT"
                                try:
                                    parts = tanggal_str.split(',')[1].strip().split()  # "17 Nov 2025 00:00:00 GMT"
                                    date_part = ' '.join(parts[:3])  # "17 Nov 2025"
                                    date_obj = datetime.datetime.strptime(date_part, '%d %b %Y')
                                except:
                                    date_obj = None
                        # Handle ISO format with time (2025-01-15T10:30:00 or 2025-01-15 10:30:00)
                        elif 'T' in tanggal_str or (' ' in tanggal_str and ':' in tanggal_str):
                            try:
                                date_obj = datetime.datetime.fromisoformat(tanggal_str.replace('Z', '+00:00'))
                            except:
                                parts = tanggal_str.split('T')[0] if 'T' in tanggal_str else tanggal_str.split(' ')[0]
                                date_obj = datetime.datetime.strptime(parts, '%Y-%m-%d')
                        # Handle date only format (2025-01-15)
                        elif '-' in tanggal_str:
                            parts = tanggal_str.split(' ')[0] if ' ' in tanggal_str else tanggal_str
                            date_obj = datetime.datetime.strptime(parts, '%Y-%m-%d')
                        # Handle already formatted dates (15/01/2025)
                        elif '/' in tanggal_str:
                            try:
                                date_obj = datetime.datetime.strptime(tanggal_str, '%d/%m/%Y')
                            except:
                                date_obj = datetime.datetime.strptime(tanggal_str, '%m/%d/%Y')
                        else:
                            date_obj = datetime.datetime.strptime(tanggal_str, '%Y-%m-%d')

                        if date_obj:
                            # Format sebagai dd/mm/yyyy saja tanpa hari
                            tanggal = date_obj.strftime('%d/%m/%Y')
                except:
                    pass
            self.table.setItem(row, 0, QTableWidgetItem(str(tanggal)))

            jenis_value = str(item.get('kategori', ''))
            jenis_item = QTableWidgetItem(jenis_value)
            jenis_item.setTextAlignment(Qt.AlignCenter)

            if jenis_value == 'Pemasukan':
                jenis_item.setBackground(QColor("#d5f4e6"))
                jenis_item.setForeground(QColor("#27ae60"))
            elif jenis_value == 'Pengeluaran':
                jenis_item.setBackground(QColor("#fadbd8"))
                jenis_item.setForeground(QColor("#c0392b"))

            jenis_item.setFlags(jenis_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 1, jenis_item)

            self.table.setItem(row, 2, QTableWidgetItem(str(item.get('sub_kategori', ''))))
            self.table.setItem(row, 3, QTableWidgetItem(str(item.get('keterangan', ''))))

            # Calculate jumlah and update running balance
            try:
                jumlah_value = float(item.get('jumlah', 0))
                if jenis_value == 'Pemasukan':
                    running_balance += jumlah_value
                else:
                    running_balance -= jumlah_value
                jumlah = f"Rp {jumlah_value:,.0f}"
            except:
                jumlah = "Rp 0"
            self.table.setItem(row, 4, QTableWidgetItem(jumlah))

            # Saldo column (running balance)
            saldo_item = QTableWidgetItem(f"Rp {running_balance:,.0f}")
            if running_balance >= 0:
                saldo_item.setForeground(QColor("#27ae60"))
            else:
                saldo_item.setForeground(QColor("#e74c3c"))
            self.table.setItem(row, 5, saldo_item)

            user_name = item.get('nama_user') or item.get('dibuat_oleh') or item.get('username') or 'System'
            self.table.setItem(row, 6, QTableWidgetItem(str(user_name)))

            tanggal_input = item.get('created_at', item.get('tanggal_input', ''))
            if tanggal_input:
                try:
                    if isinstance(tanggal_input, str):
                        if 'T' in tanggal_input or ' ' in tanggal_input:
                            date_obj = datetime.datetime.fromisoformat(tanggal_input.replace('Z', '+00:00'))
                            tanggal_input = date_obj.strftime('%d/%m/%Y %H:%M')
                        else:
                            date_obj = datetime.datetime.strptime(tanggal_input, '%Y-%m-%d')
                            tanggal_input = date_obj.strftime('%d/%m/%Y')
                except:
                    pass
            else:
                tanggal_input = '-'
            self.table.setItem(row, 7, QTableWidgetItem(str(tanggal_input)))

    def update_stats(self):
        total = len(self.filtered_data)
        pemasukan = sum(float(item.get('jumlah', 0)) for item in self.filtered_data if item.get('kategori') == 'Pemasukan')
        pengeluaran = sum(float(item.get('jumlah', 0)) for item in self.filtered_data if item.get('kategori') == 'Pengeluaran')
        saldo = pemasukan - pengeluaran

        self.total_label.setText(f"Total Transaksi: {total}")
        self.pemasukan_label.setText(f"Total Pemasukan: Rp {pemasukan:,.0f}")
        self.pengeluaran_label.setText(f"Total Pengeluaran: Rp {pengeluaran:,.0f}")
        self.saldo_label.setText(f"Saldo: Rp {saldo:,.0f}")

        if saldo >= 0:
            self.saldo_label.setStyleSheet("color: #27ae60; font-weight: bold;")
        else:
            self.saldo_label.setStyleSheet("color: #e74c3c; font-weight: bold;")


class KeuanganKategorialWidget(QWidget):
    """Tab K. Kategorial - Keuangan kategorial dikelola admin dengan CRUD"""
    log_message = pyqtSignal(str)

    def __init__(self, db_manager, admin_id=None, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.admin_id = admin_id
        self.keuangan_data = []
        self.filtered_data = []
        self.sorted_data = []  # Data sorted untuk display
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(10, 10, 10, 10)

        # Title section with professional styling
        title_frame = QFrame()
        title_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border-bottom: 2px solid #ecf0f1;
                padding: 10px 0px;
            }
        """)
        title_layout = QHBoxLayout(title_frame)
        title_layout.setContentsMargins(10, 0, 10, 0)

        title_label = QLabel("Keuangan Kategorial")
        title_font = QFont("Arial", 18, QFont.Bold)
        title_label.setFont(title_font)
        title_label.setStyleSheet("""
            QLabel {
                color: #2c3e50;
                padding: 2px;
                background-color: transparent;
                border: none;
            }
        """)
        title_layout.addWidget(title_label)
        title_layout.addStretch()
        layout.addWidget(title_frame)

        # Professional Summary Statistics
        self.create_financial_statistics_summary(layout)

        # Combined toolbar with search/filter and buttons
        toolbar_layout = QHBoxLayout()
        toolbar_layout.setSpacing(10)

        # Search
        search_label = QLabel("Cari:")
        search_label.setStyleSheet("font-weight: 500; color: #2c3e50;")
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Cari keterangan transaksi...")
        self.search_input.setFixedWidth(200)
        self.search_input.textChanged.connect(self.filter_data)

        # Type filter
        type_label = QLabel("Jenis:")
        type_label.setStyleSheet("font-weight: 500; color: #2c3e50;")
        self.jenis_filter = QComboBox()
        self.jenis_filter.addItems(["Semua", "Pemasukan", "Pengeluaran"])
        self.jenis_filter.setFixedWidth(120)
        self.jenis_filter.currentTextChanged.connect(self.filter_data)

        # Add and Refresh buttons
        add_btn = self.create_professional_button("Tambah Transaksi", "#27ae60", "#2ecc71")
        add_btn.clicked.connect(self.add_data)

        refresh_btn = self.create_professional_button("Refresh", "#3498db", "#2980b9")
        refresh_btn.clicked.connect(self.load_data)

        toolbar_layout.addWidget(search_label)
        toolbar_layout.addWidget(self.search_input)
        toolbar_layout.addWidget(type_label)
        toolbar_layout.addWidget(self.jenis_filter)
        toolbar_layout.addStretch()
        toolbar_layout.addWidget(add_btn)
        toolbar_layout.addWidget(refresh_btn)

        layout.addLayout(toolbar_layout)

        # Table with professional styling
        self.table = QTableWidget()
        self.table.setColumnCount(7)

        custom_header = WordWrapHeaderView(Qt.Horizontal, self.table)
        self.table.setHorizontalHeader(custom_header)

        self.table.setHorizontalHeaderLabels([
            "Tanggal", "Jenis", "Kategori", "Keterangan", "Jumlah (Rp)", "Saldo", "Aksi"
        ])

        self.apply_professional_table_style(self.table)

        # Set column widths
        self.table.setColumnWidth(0, 100)  # Tanggal
        self.table.setColumnWidth(1, 100)  # Jenis
        self.table.setColumnWidth(2, 120)  # Kategori
        self.table.setColumnWidth(3, 200)  # Keterangan
        self.table.setColumnWidth(4, 120)  # Jumlah
        self.table.setColumnWidth(5, 120)  # Saldo
        self.table.setColumnWidth(6, 170)  # Aksi (diperbesar untuk 3 button)

        header = self.table.horizontalHeader()
        header.sectionResized.connect(self.update_header_height)

        QTimer.singleShot(100, lambda: self.update_header_height(0, 0, 0))

        layout.addWidget(self.table)

        # Export button at bottom right
        export_layout = QHBoxLayout()
        export_layout.addStretch()

        export_excel_btn = self.create_professional_button("Export ke Excel", "#16a085", "#1abc9c")
        export_excel_btn.clicked.connect(self.export_to_excel)
        export_excel_btn.setToolTip("Ekspor data ke file Excel")
        export_layout.addWidget(export_excel_btn)

        layout.addLayout(export_layout)

    def create_financial_statistics_summary(self, layout):
        """Create financial summary using client style"""
        from PyQt5.QtWidgets import QGroupBox

        stats_group = QGroupBox("Ringkasan Keuangan")
        stats_layout = QHBoxLayout(stats_group)

        # Initialize value labels
        self.total_transaksi_value = QLabel("0")
        self.total_pemasukan_label = QLabel("Rp 0")
        self.total_pengeluaran_label = QLabel("Rp 0")
        self.total_saldo_label = QLabel("Rp 0")

        # Add stat widgets using client style
        stats_layout.addWidget(self.create_stat_widget("TOTAL TRANSAKSI", self.total_transaksi_value, "#ddeaee", "#2c3e50"))
        stats_layout.addWidget(self.create_stat_widget("TOTAL PEMASUKAN", self.total_pemasukan_label, "#e8f8f5", "#27ae60"))
        stats_layout.addWidget(self.create_stat_widget("TOTAL PENGELUARAN", self.total_pengeluaran_label, "#fadbd8", "#e74c3c"))
        stats_layout.addWidget(self.create_stat_widget("SALDO AKHIR", self.total_saldo_label, "#e3f2fd", "#3498db"))

        layout.addWidget(stats_group)

    def create_stat_widget(self, label_text, value_label, bg_color, text_color):
        """Create statistical widget using client style"""
        widget = QWidget()
        widget.setStyleSheet(f"""
            QWidget {{
                background-color: {bg_color};
                border-radius: 4px;
                padding: 6px;
            }}
        """)
        widget.setMaximumHeight(60)
        layout = QVBoxLayout(widget)
        layout.setSpacing(2)
        layout.setContentsMargins(6, 6, 6, 6)

        label = QLabel(label_text)
        label.setAlignment(Qt.AlignCenter)
        label.setWordWrap(True)
        label.setStyleSheet(f"""
            QLabel {{
                font-weight: bold;
                color: {text_color};
                font-size: 9px;
                font-family: Arial, sans-serif;
                background-color: transparent;
                min-height: 12px;
            }}
        """)

        value_label.setAlignment(Qt.AlignCenter)
        value_label.setWordWrap(True)
        value_label.setStyleSheet(f"""
            QLabel {{
                font-size: 14px;
                font-weight: bold;
                color: {text_color};
                font-family: Arial, sans-serif;
                background-color: transparent;
                padding: 2px;
                min-height: 16px;
            }}
        """)

        layout.addWidget(label)
        layout.addWidget(value_label)

        return widget

    def create_professional_button(self, text, bg_color, hover_color):
        """Create a professional button with gradient"""
        button = QPushButton(text)
        button.setMinimumSize(150, 32)
        button.setMaximumSize(220, 32)

        button.setStyleSheet(f"""
            QPushButton {{
                background-color: {bg_color};
                color: white;
                border: 1px solid {bg_color};
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 11px;
                font-weight: 500;
                text-align: center;
                margin: 1px 3px;
                box-shadow: 0 1px 2px rgba(0,0,0,0.1);
            }}
            QPushButton:hover {{
                background-color: {hover_color};
                border-color: {hover_color};
                box-shadow: 0 2px 4px rgba(0,0,0,0.15);
            }}
            QPushButton:pressed {{
                background-color: {bg_color};
                border-color: {bg_color};
                box-shadow: inset 0 1px 2px rgba(0,0,0,0.2);
            }}
        """)

        return button

    def update_header_height(self, logicalIndex, oldSize, newSize):
        if hasattr(self, 'table'):
            header = self.table.horizontalHeader()
            header.setMinimumHeight(25)
            max_height = 25
            for i in range(header.count()):
                size = header.sectionSizeFromContents(i)
                max_height = max(max_height, size.height())
            header.setFixedHeight(max_height)

    def apply_professional_table_style(self, table):
        header_font = QFont()
        header_font.setBold(True)
        header_font.setPointSize(9)
        table.horizontalHeader().setFont(header_font)

        table.horizontalHeader().setStyleSheet("""
            QHeaderView::section {
                background-color: #f2f2f2;
                border: none;
                border-bottom: 1px solid #d4d4d4;
                border-right: 1px solid #d4d4d4;
                padding: 6px 4px;
                font-weight: bold;
                color: #333333;
            }
        """)

        header = table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignCenter | Qt.AlignVCenter)
        header.setSectionsClickable(True)
        header.setMinimumHeight(25)
        header.setMinimumSectionSize(50)
        header.setMaximumSectionSize(500)
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setStretchLastSection(True)

        table.setStyleSheet("""
            QTableWidget {
                gridline-color: #d4d4d4;
                background-color: white;
                border: 1px solid #d4d4d4;
                selection-background-color: #cce7ff;
                font-family: 'Calibri', 'Segoe UI', Arial, sans-serif;
                font-size: 9pt;
                outline: none;
            }
            QTableWidget::item {
                border: none;
                padding: 4px 6px;
                min-height: 18px;
            }
            QTableWidget::item:selected {
                background-color: #cce7ff;
                color: black;
            }
        """)

        table.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)

        table.verticalHeader().setDefaultSectionSize(20)
        table.setSelectionBehavior(QAbstractItemView.SelectItems)
        table.setAlternatingRowColors(False)
        table.verticalHeader().setVisible(True)
        table.setShowGrid(True)
        table.setGridStyle(Qt.SolidLine)
        table.setMinimumHeight(150)

    def load_data(self):
        try:
            print(f"[DEBUG] KeuanganKategorialWidget.load_data - loading...")
            success, data = self.db_manager.get_keuangan_kategorial_list()
            print(f"[DEBUG] KeuanganKategorialWidget.load_data - success={success}, data type={type(data)}")
            if success:
                self.keuangan_data = data if isinstance(data, list) else []
                self.filtered_data = self.keuangan_data.copy()
                self.update_table()
                self.update_stats()
                self.log_message.emit(f"Data keuangan kategorial dimuat: {len(self.keuangan_data)} record")
                print(f"[DEBUG] KeuanganKategorialWidget.load_data - loaded {len(self.keuangan_data)} records")
            else:
                print(f"[ERROR] KeuanganKategorialWidget.load_data failed: {data}")
                self.log_message.emit(f"Gagal memuat data: {data}")
        except Exception as e:
            print(f"[ERROR] KeuanganKategorialWidget.load_data exception: {str(e)}")
            import traceback
            traceback.print_exc()
            self.log_message.emit(f"Error: {str(e)}")

    def filter_data(self):
        search = self.search_input.text().lower()
        jenis = self.jenis_filter.currentText()

        self.filtered_data = []
        for item in self.keuangan_data:
            if search and search not in str(item.get('keterangan', '')).lower():
                continue

            if jenis != "Semua" and item.get('jenis') != jenis:
                continue

            self.filtered_data.append(item)

        self.update_table()
        self.update_stats()

    def update_table(self):
        # Sort data by ID (oldest first) for correct running balance
        self.sorted_data = sorted(self.filtered_data, key=lambda x: x.get('id_keuangan_kategorial', 0) or x.get('id', 0), reverse=False)

        self.table.setRowCount(len(self.sorted_data))

        # Calculate running balance
        running_balance = 0.0

        for row, item in enumerate(self.sorted_data):
            # Set row height
            self.table.setRowHeight(row, 36)

            # Format tanggal transaksi - dd/mm/yyyy saja tanpa hari
            tanggal = item.get('tanggal', '')
            if tanggal:
                try:
                    if isinstance(tanggal, str):
                        tanggal_str = tanggal.strip()

                        # Handle RFC 1123 / GMT format (Mon, 17 Nov 2025 00:00:00 GMT)
                        if 'GMT' in tanggal_str or tanggal_str.count(',') == 1:
                            try:
                                # Parse RFC 1123 format
                                from email.utils import parsedate_to_datetime
                                date_obj = parsedate_to_datetime(tanggal_str)
                            except:
                                # Fallback: extract date parts manually
                                # Format: "Mon, 17 Nov 2025 00:00:00 GMT"
                                try:
                                    parts = tanggal_str.split(',')[1].strip().split()  # "17 Nov 2025 00:00:00 GMT"
                                    date_part = ' '.join(parts[:3])  # "17 Nov 2025"
                                    date_obj = datetime.datetime.strptime(date_part, '%d %b %Y')
                                except:
                                    date_obj = None
                        # Handle ISO format with time (2025-01-15T10:30:00 or 2025-01-15 10:30:00)
                        elif 'T' in tanggal_str or (' ' in tanggal_str and ':' in tanggal_str):
                            try:
                                date_obj = datetime.datetime.fromisoformat(tanggal_str.replace('Z', '+00:00'))
                            except:
                                parts = tanggal_str.split('T')[0] if 'T' in tanggal_str else tanggal_str.split(' ')[0]
                                date_obj = datetime.datetime.strptime(parts, '%Y-%m-%d')
                        # Handle date only format (2025-01-15)
                        elif '-' in tanggal_str:
                            parts = tanggal_str.split(' ')[0] if ' ' in tanggal_str else tanggal_str
                            date_obj = datetime.datetime.strptime(parts, '%Y-%m-%d')
                        # Handle already formatted dates (15/01/2025)
                        elif '/' in tanggal_str:
                            try:
                                date_obj = datetime.datetime.strptime(tanggal_str, '%d/%m/%Y')
                            except:
                                date_obj = datetime.datetime.strptime(tanggal_str, '%m/%d/%Y')
                        else:
                            date_obj = datetime.datetime.strptime(tanggal_str, '%Y-%m-%d')

                        if date_obj:
                            # Format sebagai dd/mm/yyyy saja tanpa hari
                            tanggal = date_obj.strftime('%d/%m/%Y')
                except:
                    pass

            tanggal_item = QTableWidgetItem(str(tanggal))
            tanggal_item.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(row, 0, tanggal_item)

            # Jenis (Pemasukan/Pengeluaran) - dari field 'jenis'
            jenis_value = str(item.get('jenis', ''))
            jenis_item = QTableWidgetItem(jenis_value)
            jenis_item.setTextAlignment(Qt.AlignCenter)

            if jenis_value == 'Pemasukan':
                jenis_item.setBackground(QColor("#d5f4e6"))
                jenis_item.setForeground(QColor("#27ae60"))
            elif jenis_value == 'Pengeluaran':
                jenis_item.setBackground(QColor("#fadbd8"))
                jenis_item.setForeground(QColor("#c0392b"))

            jenis_item.setFlags(jenis_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 1, jenis_item)

            # Kategori (Kolekte Mingguan, Donasi, etc) - dari field 'kategori'
            kat_item = QTableWidgetItem(str(item.get('kategori', '')))
            kat_item.setFlags(kat_item.flags() & ~Qt.ItemIsEditable)
            self.table.setItem(row, 2, kat_item)

            self.table.setItem(row, 3, QTableWidgetItem(str(item.get('keterangan', ''))))

            # Calculate running balance and jumlah
            try:
                jumlah_value = float(item.get('jumlah', 0))
                if jenis_value == 'Pemasukan':
                    running_balance += jumlah_value
                    jumlah_text = f"+{jumlah_value:,.0f}"
                    jumlah_color = QColor("#27ae60")
                else:
                    running_balance -= jumlah_value
                    jumlah_text = f"-{jumlah_value:,.0f}"
                    jumlah_color = QColor("#e74c3c")
            except:
                jumlah_text = "0"
                jumlah_color = QColor("#333333")

            jumlah_item = QTableWidgetItem(jumlah_text)
            jumlah_item.setForeground(jumlah_color)
            self.table.setItem(row, 4, jumlah_item)

            # Saldo column (running balance)
            saldo_item = QTableWidgetItem(f"{running_balance:,.0f}")
            if running_balance >= 0:
                saldo_item.setForeground(QColor("#27ae60"))
            else:
                saldo_item.setForeground(QColor("#e74c3c"))
            self.table.setItem(row, 5, saldo_item)

            # Aksi buttons - professional style matching client
            action_widget = self.create_action_buttons_for_row(row)
            self.table.setCellWidget(row, 6, action_widget)

    def create_action_buttons_for_row(self, row):
        """Create action buttons for table row - matching client style with 3 buttons"""
        action_widget = QWidget()
        action_layout = QHBoxLayout(action_widget)
        action_layout.setContentsMargins(3, 3, 3, 3)
        action_layout.setSpacing(3)
        action_layout.setAlignment(Qt.AlignCenter)

        # View button
        view_button = QPushButton()
        view_button.setText("Lihat")
        view_button.setFixedSize(50, 28)
        view_button.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 2px;
                font-size: 10px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
            QPushButton:pressed {
                background-color: #21618c;
            }
        """)
        view_button.setToolTip("Lihat Detail Transaksi")
        view_button.clicked.connect(lambda _, r=row: self.view_data(r))
        action_layout.addWidget(view_button)

        # Edit button
        edit_button = QPushButton()
        edit_button.setText("Edit")
        edit_button.setFixedSize(50, 28)
        edit_button.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 2px;
                font-size: 10px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #e67e22;
            }
            QPushButton:pressed {
                background-color: #d35400;
            }
        """)
        edit_button.setToolTip("Edit Transaksi")
        edit_button.clicked.connect(lambda _, r=row: self.edit_data(r))
        action_layout.addWidget(edit_button)

        # Delete button
        delete_button = QPushButton()
        delete_button.setText("Hapus")
        delete_button.setFixedSize(55, 28)
        delete_button.setStyleSheet("""
            QPushButton {
                background-color: #e74c3c;
                color: white;
                border: none;
                border-radius: 3px;
                padding: 2px;
                font-size: 10px;
                font-weight: 500;
            }
            QPushButton:hover {
                background-color: #c0392b;
            }
            QPushButton:pressed {
                background-color: #a93226;
            }
        """)
        delete_button.setToolTip("Hapus Transaksi")
        delete_button.clicked.connect(lambda _, r=row: self.delete_data(r))
        action_layout.addWidget(delete_button)

        return action_widget

    def update_stats(self):
        """Update statistics labels using client style"""
        total = len(self.filtered_data)
        pemasukan = sum(float(item.get('jumlah', 0)) for item in self.filtered_data if item.get('jenis') == 'Pemasukan')
        pengeluaran = sum(float(item.get('jumlah', 0)) for item in self.filtered_data if item.get('jenis') == 'Pengeluaran')
        saldo = pemasukan - pengeluaran

        # Update summary widgets using client style
        if hasattr(self, 'total_transaksi_value'):
            self.total_transaksi_value.setText(str(total))

        if hasattr(self, 'total_pemasukan_label'):
            self.total_pemasukan_label.setText(f"Rp {pemasukan:,.0f}")

        if hasattr(self, 'total_pengeluaran_label'):
            self.total_pengeluaran_label.setText(f"Rp {pengeluaran:,.0f}")

        if hasattr(self, 'total_saldo_label'):
            self.total_saldo_label.setText(f"Rp {saldo:,.0f}")
            # Update saldo color dynamically
            color = "#3498db" if saldo >= 0 else "#e74c3c"
            bg_color = "#e3f2fd" if saldo >= 0 else "#fadbd8"

            # Update parent widget background color too
            if self.total_saldo_label.parent():
                self.total_saldo_label.parent().setStyleSheet(f"""
                    QWidget {{
                        background-color: {bg_color};
                        border-radius: 4px;
                        padding: 6px;
                    }}
                """)

            self.total_saldo_label.setStyleSheet(f"""
                QLabel {{
                    font-size: 14px;
                    font-weight: bold;
                    color: {color};
                    font-family: Arial, sans-serif;
                    background-color: transparent;
                    padding: 2px;
                    min-height: 16px;
                }}
            """)

    def add_data(self):
        dialog = KeuanganDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            print(f"[DEBUG] add_data - data being sent: {data}, admin_id: {self.admin_id}")
            try:
                success, result = self.db_manager.add_keuangan_kategorial(data, self.admin_id)
                print(f"[DEBUG] add_data - API response: success={success}, result={result}")
                if success:
                    self.load_data()
                    self.log_message.emit(f"Data keuangan berhasil ditambahkan")
                else:
                    print(f"[ERROR] Add keuangan kategorial failed: {result}")
                    QMessageBox.critical(self, "Error", f"Gagal menambah data:\n{result}")
            except Exception as e:
                print(f"[ERROR] Exception in add_data: {str(e)}")
                import traceback
                traceback.print_exc()
                QMessageBox.critical(self, "Error", f"Error:\n{str(e)}")

    def set_admin_id(self, admin_id):
        """Set admin ID for this widget"""
        self.admin_id = admin_id

    def view_data(self, row):
        """View transaction detail from specific row - using sorted_data"""
        try:
            if row < 0 or row >= len(self.sorted_data):
                print(f"[ERROR] Invalid row index: {row}, sorted_data length: {len(self.sorted_data)}")
                return

            item = self.sorted_data[row]
            print(f"[DEBUG] view_data - viewing row={row}, data: {item}")

            # Format tanggal - dd/mm/yyyy
            tanggal = item.get('tanggal', 'N/A')
            if tanggal and tanggal != 'N/A':
                try:
                    if isinstance(tanggal, str):
                        tanggal_str = tanggal.strip()
                        if 'T' in tanggal_str or (' ' in tanggal_str and ':' in tanggal_str):
                            try:
                                dt = datetime.datetime.fromisoformat(tanggal_str.replace('Z', '+00:00'))
                                tanggal = dt.strftime('%d/%m/%Y')
                            except:
                                parts = tanggal_str.split(' ')[0] if ' ' in tanggal_str else tanggal_str
                                dt = datetime.datetime.strptime(parts, '%Y-%m-%d')
                                tanggal = dt.strftime('%d/%m/%Y')
                        elif '-' in tanggal_str:
                            parts = tanggal_str.split(' ')[0]
                            dt = datetime.datetime.strptime(parts, '%Y-%m-%d')
                            tanggal = dt.strftime('%d/%m/%Y')
                        elif '/' in tanggal_str:
                            # Already formatted
                            pass
                except Exception as e:
                    print(f"[WARN] Failed to format date: {e}")

            # Get other details
            jenis = item.get('jenis', 'N/A')
            kategori = item.get('kategori', 'N/A')
            keterangan = item.get('keterangan', 'N/A')
            jumlah = item.get('jumlah', 0)

            detail_text = f"""
Detail Transaksi Keuangan Kategorial:

Tanggal: {tanggal}
Jenis: {jenis}
Kategori: {kategori}
Keterangan: {keterangan}
Jumlah: Rp {jumlah:,.0f}
"""

            QMessageBox.information(self, "Detail Transaksi", detail_text)
        except Exception as e:
            print(f"[ERROR] Exception in view_data: {str(e)}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Error menampilkan detail: {str(e)}")

    def edit_data(self, row):
        """Edit transaction from specific row - using sorted_data"""
        try:
            if row < 0 or row >= len(self.sorted_data):
                print(f"[ERROR] Invalid row index: {row}, sorted_data length: {len(self.sorted_data)}")
                return

            item = self.sorted_data[row]
            item_id = item.get('id_keuangan_kategorial')
            print(f"[DEBUG] edit_data - editing row={row}, ID: {item_id}")

            dialog = KeuanganDialog(self, item)
            if dialog.exec_() == QDialog.Accepted:
                data = dialog.get_data()
                print(f"[DEBUG] edit_data - updated data: {data}")
                try:
                    success, result = self.db_manager.update_keuangan_kategorial(item_id, data)
                    print(f"[DEBUG] edit_data - API response: success={success}, result={result}")
                    if success:
                        self.load_data()
                        self.log_message.emit(f"Data keuangan berhasil diupdate")
                        QMessageBox.information(self, "Sukses", "Transaksi berhasil diperbarui")
                    else:
                        print(f"[ERROR] Update keuangan kategorial failed: {result}")
                        QMessageBox.critical(self, "Error", f"Gagal update data:\n{result}")
                except Exception as e:
                    print(f"[ERROR] Exception in edit_data: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    QMessageBox.critical(self, "Error", f"Error:\n{str(e)}")
        except Exception as e:
            print(f"[ERROR] Exception in edit_data: {str(e)}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Error mengedit transaksi: {str(e)}")

    def delete_data(self, row):
        """Delete transaction from specific row - using sorted_data"""
        try:
            if row < 0 or row >= len(self.sorted_data):
                print(f"[ERROR] Invalid row index: {row}, sorted_data length: {len(self.sorted_data)}")
                return

            item = self.sorted_data[row]
            item_id = item.get('id_keuangan_kategorial')
            keterangan = item.get('keterangan', 'Unknown')

            reply = QMessageBox.question(self, "Konfirmasi",
                                       f"Apakah Anda yakin ingin menghapus transaksi '{keterangan}'?",
                                       QMessageBox.Yes | QMessageBox.No)

            if reply == QMessageBox.Yes:
                print(f"[DEBUG] delete_data - deleting row={row}, ID: {item_id}")
                try:
                    success, result = self.db_manager.delete_keuangan_kategorial(item_id)
                    print(f"[DEBUG] delete_data - API response: success={success}, result={result}")
                    if success:
                        self.load_data()
                        self.log_message.emit(f"Data keuangan berhasil dihapus")
                        QMessageBox.information(self, "Sukses", "Transaksi berhasil dihapus")
                    else:
                        print(f"[ERROR] Delete keuangan kategorial failed: {result}")
                        QMessageBox.critical(self, "Error", f"Gagal hapus data:\n{result}")
                except Exception as e:
                    print(f"[ERROR] Exception in delete_data: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    QMessageBox.critical(self, "Error", f"Error:\n{str(e)}")
        except Exception as e:
            print(f"[ERROR] Exception in delete_data: {str(e)}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Error menghapus transaksi: {str(e)}")

    def export_to_excel(self):
        """Export keuangan kategorial data to Excel file"""
        if not self.filtered_data:
            QMessageBox.warning(self, "Warning", "Tidak ada data untuk diekspor")
            return

        try:
            # Get save file path
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Export Data Keuangan Kategorial",
                f"keuangan_kategorial_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if not file_path:
                return

            # Try to import openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            except ImportError:
                QMessageBox.critical(
                    self,
                    "Error",
                    "Library openpyxl tidak ditemukan.\nInstall dengan: pip install openpyxl"
                )
                return

            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Keuangan Kategorial"

            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Write headers
            headers = ["No", "Tanggal", "Jenis", "Kategori", "Keterangan", "Jumlah (Rp)", "Saldo"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = str(header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_style

            # Set column widths
            ws.column_dimensions['A'].width = 5  # type: ignore
            ws.column_dimensions['B'].width = 12  # type: ignore
            ws.column_dimensions['C'].width = 12  # type: ignore
            ws.column_dimensions['D'].width = 20  # type: ignore
            ws.column_dimensions['E'].width = 35  # type: ignore
            ws.column_dimensions['F'].width = 15  # type: ignore
            ws.column_dimensions['G'].width = 15  # type: ignore

            # Sort data by ID for correct running balance
            sorted_data = sorted(self.filtered_data, key=lambda x: x.get('id_keuangan_kategorial', 0) or x.get('id', 0), reverse=False)

            # Calculate running balance and write data
            running_balance = 0.0

            for idx, item in enumerate(sorted_data, 1):
                # Format tanggal
                tanggal = item.get('tanggal', '')
                if tanggal:
                    try:
                        if isinstance(tanggal, str):
                            tanggal_str = tanggal.strip()
                            if 'T' in tanggal_str or (' ' in tanggal_str and ':' in tanggal_str):
                                dt = datetime.datetime.fromisoformat(tanggal_str.replace('Z', '+00:00'))
                                tanggal = dt.strftime('%d/%m/%Y')
                            elif '-' in tanggal_str:
                                parts = tanggal_str.split(' ')[0]
                                dt = datetime.datetime.strptime(parts, '%Y-%m-%d')
                                tanggal = dt.strftime('%d/%m/%Y')
                    except:
                        pass

                jenis = item.get('jenis', '')
                kategori = item.get('kategori', '')
                keterangan = item.get('keterangan', '')

                try:
                    jumlah_value = float(item.get('jumlah', 0))
                    if jenis == 'Pemasukan':
                        running_balance += jumlah_value
                    else:
                        running_balance -= jumlah_value
                except:
                    jumlah_value = 0

                # Write row data
                row_num = idx + 1

                # Column 1: No
                cell_no = ws.cell(row=row_num, column=1)
                cell_no.value = idx
                cell_no.border = border_style

                # Column 2: Tanggal
                cell_tanggal = ws.cell(row=row_num, column=2)
                cell_tanggal.value = str(tanggal)
                cell_tanggal.border = border_style

                # Column 3: Jenis
                jenis_cell = ws.cell(row=row_num, column=3)
                jenis_cell.value = str(jenis)
                jenis_cell.border = border_style
                if jenis == 'Pemasukan':
                    jenis_cell.fill = PatternFill(start_color="d5f4e6", end_color="d5f4e6", fill_type="solid")
                    jenis_cell.font = Font(color="27ae60", bold=True)
                elif jenis == 'Pengeluaran':
                    jenis_cell.fill = PatternFill(start_color="fadbd8", end_color="fadbd8", fill_type="solid")
                    jenis_cell.font = Font(color="c0392b", bold=True)

                # Column 4: Kategori
                cell_kategori = ws.cell(row=row_num, column=4)
                cell_kategori.value = str(kategori)
                cell_kategori.border = border_style

                # Column 5: Keterangan
                cell_keterangan = ws.cell(row=row_num, column=5)
                cell_keterangan.value = str(keterangan)
                cell_keterangan.border = border_style

                # Column 6: Jumlah
                jumlah_cell = ws.cell(row=row_num, column=6)
                jumlah_cell.value = jumlah_value
                jumlah_cell.border = border_style
                jumlah_cell.number_format = '#,##0'
                if jenis == 'Pemasukan':
                    jumlah_cell.font = Font(color="27ae60")
                else:
                    jumlah_cell.font = Font(color="c0392b")

                # Column 7: Saldo
                saldo_cell = ws.cell(row=row_num, column=7)
                saldo_cell.value = running_balance
                saldo_cell.border = border_style
                saldo_cell.number_format = '#,##0'
                if running_balance >= 0:
                    saldo_cell.font = Font(color="27ae60", bold=True)
                else:
                    saldo_cell.font = Font(color="c0392b", bold=True)

            # Add summary at the bottom
            summary_row = len(sorted_data) + 3

            # Calculate totals
            total_pemasukan = sum(float(item.get('jumlah', 0)) for item in sorted_data if item.get('jenis') == 'Pemasukan')
            total_pengeluaran = sum(float(item.get('jumlah', 0)) for item in sorted_data if item.get('jenis') == 'Pengeluaran')
            saldo_akhir = total_pemasukan - total_pengeluaran

            # Summary header
            summary_header = ws.cell(row=summary_row, column=4)
            summary_header.value = "RINGKASAN:"
            summary_header.font = Font(bold=True)

            # Total Pemasukan
            label_pemasukan = ws.cell(row=summary_row + 1, column=4)
            label_pemasukan.value = "Total Pemasukan:"

            pemasukan_cell = ws.cell(row=summary_row + 1, column=5)
            pemasukan_cell.value = total_pemasukan
            pemasukan_cell.number_format = '#,##0'
            pemasukan_cell.font = Font(color="27ae60", bold=True)

            # Total Pengeluaran
            label_pengeluaran = ws.cell(row=summary_row + 2, column=4)
            label_pengeluaran.value = "Total Pengeluaran:"

            pengeluaran_cell = ws.cell(row=summary_row + 2, column=5)
            pengeluaran_cell.value = total_pengeluaran
            pengeluaran_cell.number_format = '#,##0'
            pengeluaran_cell.font = Font(color="c0392b", bold=True)

            # Saldo Akhir
            label_saldo = ws.cell(row=summary_row + 3, column=4)
            label_saldo.value = "Saldo Akhir:"

            saldo_cell_summary = ws.cell(row=summary_row + 3, column=5)
            saldo_cell_summary.value = saldo_akhir
            saldo_cell_summary.number_format = '#,##0'
            saldo_cell_summary.font = Font(color="27ae60" if saldo_akhir >= 0 else "c0392b", bold=True)

            # Save workbook
            wb.save(file_path)

            QMessageBox.information(
                self,
                "Export Berhasil",
                f"Data berhasil diekspor ke:\n{file_path}\n\nTotal: {len(sorted_data)} transaksi"
            )
            self.log_message.emit(f"Data keuangan kategorial berhasil diekspor ke {file_path}")

        except Exception as e:
            print(f"[ERROR] Exception in export_to_excel: {str(e)}")
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Gagal mengekspor data:\n{str(e)}")
            self.log_message.emit(f"Error export data: {str(e)}")


class KeuanganComponent(QWidget):
    """Main Keuangan Component dengan tab WR dan K. Kategorial"""
    log_message = pyqtSignal(str)

    def __init__(self, db_manager, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.current_admin = None

        layout = QVBoxLayout(self)

        # Tab widget
        self.tab_widget = QTabWidget()

        # Tab WR
        self.wr_widget = KeuanganWRWidget(db_manager, self)
        self.wr_widget.log_message.connect(self.on_log_message)
        self.tab_widget.addTab(self.wr_widget, "WR")

        # Tab K. Kategorial
        self.kategorial_widget = KeuanganKategorialWidget(db_manager, None, self)
        self.kategorial_widget.log_message.connect(self.on_log_message)
        self.tab_widget.addTab(self.kategorial_widget, "K. Kategorial")

        layout.addWidget(self.tab_widget)

    def load_data(self):
        """Load data in both tabs"""
        try:
            self.wr_widget.load_data()
            self.kategorial_widget.load_data()
            self.log_message.emit("Data keuangan dimuat")
        except Exception as e:
            self.log_message.emit(f"Error loading data: {str(e)}")

    def set_current_admin(self, admin_data):
        """Set the current admin data."""
        self.current_admin = admin_data
        # Propagate admin_id to kategorial widget
        if admin_data and 'id_admin' in admin_data:
            self.kategorial_widget.set_admin_id(admin_data['id_admin'])

    def on_log_message(self, message):
        """Relay log messages"""
        self.log_message.emit(message)
